"""ref_reader.readers — 1차 포맷 리더 (primary readers).

포맷:
- text (.txt, .md, .log)
- csv (.csv)
- excel (.xlsx, .xls)
- json (.json)
- pdf (.pdf)
- html (.html, .htm)
- xml (.xml)
- hwp_structured (.hwp, .hwpx via hwp_analyzer.analyze_document)

변환 기반 리더 (docx, pptx, rtf 등) 는 conversion.py 에 있음.
"""
import os
import sys
import json
import re


def _read_text(path, max_chars):
    with open(path, 'r', encoding='utf-8', errors='replace') as f:
        content = f.read(max_chars)
    return {
        "format": "text",
        "file_name": os.path.basename(path),
        "content": content,
        "char_count": len(content),
    }


def _read_csv(path, max_chars):
    import csv
    rows = []
    total_chars = 0
    with open(path, 'r', encoding='utf-8', errors='replace') as f:
        reader = csv.reader(f)
        for row in reader:
            row_text = ','.join(row)
            total_chars += len(row_text)
            if total_chars > max_chars:
                break
            rows.append(row)

    headers = rows[0] if rows else []
    data = rows[1:] if len(rows) > 1 else []
    return {
        "format": "csv",
        "file_name": os.path.basename(path),
        "headers": headers,
        "data": data,
        "row_count": len(data),
    }


def _read_excel(path, max_chars):
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl이 필요합니다. pip install openpyxl")

    wb = None
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            total_chars = 0
            for row in ws.iter_rows(values_only=True):
                row_data = [str(cell) if cell is not None else "" for cell in row]
                total_chars += sum(len(c) for c in row_data)
                if total_chars > max_chars:
                    break
                rows.append(row_data)

            headers = rows[0] if rows else []
            data = rows[1:] if len(rows) > 1 else []
            sheets.append({
                "sheet_name": sheet_name,
                "headers": headers,
                "data": data,
                "row_count": len(data),
            })

        return {
            "format": "excel",
            "file_name": os.path.basename(path),
            "sheets": sheets,
            "sheet_count": len(sheets),
        }
    finally:
        if wb:
            wb.close()


def _read_json(path, max_chars):
    with open(path, 'r', encoding='utf-8', errors='replace') as f:
        content = f.read(max_chars)

    data = json.loads(content)
    return {
        "format": "json",
        "file_name": os.path.basename(path),
        "data": data,
    }


def _read_pdf(path, max_chars):
    """PDF에서 텍스트 추출 (PyMuPDF 사용)."""
    try:
        import fitz  # PyMuPDF
    except ImportError:
        raise ImportError("PyMuPDF가 필요합니다. pip install PyMuPDF")

    doc = fitz.open(path)
    pages = []
    total_chars = 0
    for i, page in enumerate(doc):
        text = page.get_text("text")
        total_chars += len(text)
        pages.append({"page": i + 1, "text": text})
        if total_chars > max_chars:
            break
    doc.close()

    full_text = "\n\n".join(p["text"] for p in pages)
    return {
        "format": "pdf",
        "file_name": os.path.basename(path),
        "content": full_text[:max_chars],
        "page_count": len(pages),
        "char_count": len(full_text[:max_chars]),
    }


def _read_html(path, max_chars):
    """HTML 파일에서 텍스트 + 표 추출."""
    with open(path, 'r', encoding='utf-8', errors='replace') as f:
        html = f.read(max_chars * 3)  # HTML 태그 포함이므로 더 많이 읽음

    # 표 추출 (정규식 기반 — BeautifulSoup 없이)
    tables = []
    table_pattern = re.compile(r'<table[^>]*>(.*?)</table>', re.DOTALL | re.IGNORECASE)
    for match in table_pattern.finditer(html):
        table_html = match.group(1)
        rows = []
        row_pattern = re.compile(r'<tr[^>]*>(.*?)</tr>', re.DOTALL | re.IGNORECASE)
        for row_match in row_pattern.finditer(table_html):
            row_html = row_match.group(1)
            cells = re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>', row_html, re.DOTALL | re.IGNORECASE)
            # 태그 제거
            clean_cells = [re.sub(r'<[^>]+>', '', c).strip() for c in cells]
            if clean_cells:
                rows.append(clean_cells)
        if rows:
            headers = rows[0] if rows else []
            data = rows[1:] if len(rows) > 1 else []
            tables.append({"headers": headers, "data": data, "row_count": len(data)})

    # 본문 텍스트 (태그 제거)
    text = re.sub(r'<style[^>]*>.*?</style>', '', html, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r'<script[^>]*>.*?</script>', '', text, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r'<[^>]+>', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()[:max_chars]

    return {
        "format": "html",
        "file_name": os.path.basename(path),
        "content": text,
        "tables": tables,
        "table_count": len(tables),
        "char_count": len(text),
    }


def _read_xml(path, max_chars):
    """XML 파일에서 구조화된 데이터 추출."""
    import xml.etree.ElementTree as ET

    tree = ET.parse(path)
    root = tree.getroot()

    # 네임스페이스 제거 (간편 접근)
    def strip_ns(tag):
        return tag.split('}')[-1] if '}' in tag else tag

    def elem_to_dict(elem):
        """XML 요소를 dict로 변환 (재귀)."""
        result = {}
        # 속성
        if elem.attrib:
            result["@attributes"] = dict(elem.attrib)
        # 텍스트
        if elem.text and elem.text.strip():
            result["@text"] = elem.text.strip()
        # 자식 요소
        children = {}
        for child in elem:
            tag = strip_ns(child.tag)
            child_data = elem_to_dict(child)
            if tag in children:
                # 같은 태그 반복 → 리스트화
                if not isinstance(children[tag], list):
                    children[tag] = [children[tag]]
                children[tag].append(child_data)
            else:
                children[tag] = child_data
        result.update(children)
        return result

    data = {strip_ns(root.tag): elem_to_dict(root)}

    # 텍스트 미리보기
    text_parts = []
    for elem in root.iter():
        if elem.text and elem.text.strip():
            text_parts.append(elem.text.strip())
    content = "\n".join(text_parts)[:max_chars]

    return {
        "format": "xml",
        "file_name": os.path.basename(path),
        "data": data,
        "content": content,
        "char_count": len(content),
    }


def _read_hwp_structured(hwp, file_path, max_chars):
    """v0.7.4.8 Fix C1: HWP 파일을 구조화된 tables + full_text 로 반환.

    기존 PDF 변환 경로와 달리 표 구조(rows × cols)를 2D 배열로 보존.
    hwp_analyzer.analyze_document 가 이미 hwp.table_to_df() 로 2D 추출 중이므로 재사용.
    """
    try:
        from hwp_analyzer import analyze_document
    except ImportError as e:
        raise ValueError(f"hwp_analyzer import 실패: {e}") from e

    try:
        analysis = analyze_document(hwp, file_path, already_open=False)
    except Exception as e:
        # 분석 실패 시 PDF 변환 fallback (구조 손실되지만 텍스트는 나옴)
        print(f"[WARN] _read_hwp_structured analyze failed: {e} — PDF fallback",
              file=sys.stderr)
        # lazy import to avoid circular (conversion → readers._read_pdf)
        from .conversion import _read_via_pdf_conversion
        return _read_via_pdf_conversion(file_path, max_chars)

    # analyze_document 반환 구조: {tables: [...], fields: [...], text_preview, full_text, ...}
    full_text = analysis.get("full_text", "") or ""
    original_len = len(full_text)
    truncated = False
    if original_len > max_chars:
        full_text = full_text[:max_chars]
        truncated = True

    # tables 를 ref_reader 표준 schema 로 변환
    tables_out = []
    for tbl in analysis.get("tables", []) or []:
        headers = tbl.get("headers", []) or []
        data = tbl.get("data", []) or []
        tables_out.append({
            "index": tbl.get("index", 0),
            "rows": tbl.get("rows", len(data) + (1 if headers else 0)),
            "cols": tbl.get("cols", len(headers) if headers else (len(data[0]) if data else 0)),
            "headers": headers,
            "data": data,
        })

    # 기존 csv/excel 경로와 호환되는 최상위 headers/data 도 제공
    primary_headers: list = []
    primary_data: list = []
    if tables_out:
        first = tables_out[0]
        primary_headers = first.get("headers", []) or []
        primary_data = first.get("data", []) or []

    result = {
        "format": "hwp_structured",
        "file_name": os.path.basename(file_path),
        "file_format": "HWPX" if file_path.lower().endswith(".hwpx") else "HWP",
        "tables": tables_out,
        "table_count": len(tables_out),
        "fields": analysis.get("fields", []) or [],
        "full_text": full_text,
        "text_preview": analysis.get("text_preview", "") or "",
        "pages": analysis.get("pages", 0) or 0,
        "char_count": len(full_text),
        "original_char_count": original_len,
        "truncated": truncated,
        # 호환 필드 (기존 csv/excel 경로)
        "headers": primary_headers,
        "data": primary_data,
    }
    if truncated:
        result["warning"] = (
            f"full_text 가 {original_len} 글자 → {max_chars} 글자로 잘렸습니다. "
            f"max_chars 를 더 크게 설정하거나 full_text 대신 tables[] 를 사용하세요."
        )
    return result
